"""
🤖 GROK AUTO GENERATOR BOT — 24-Jam Non-Stop Generate
Terus-menerus generate video Grok untuk mengisi stok:
  1. Brutal Bot (brutal_stok/)
  2. Tiap UD di Grok TikTok Bot (gtt_stok/ud_X/)
  3. Stok Global (global_stok/) jika semua sudah penuh

Fitur:
  - Prioritas stok yang bisa diatur (default: Brutal → UD1 → UD2 → ...)
  - Spreadsheet logging (nama stok, datetime, status)
  - Kelola bahan & prompt per target
  - Ambil stok global → kirim video ke Telegram
  - Cek stok global dulu saat Generate/Stok Sekarang
  - Retry terus jika gagal
"""

import os, sys, re, time, shutil, asyncio, json, threading, random, glob, logging, copy
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

from telegram import Update, BotCommand, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes, ConversationHandler
)
from telegram.constants import ParseMode

sys.path.insert(0, r"c:\tiktok_automation")
from gtt_core import (
    APP_DIR, BAHAN_DIR, DB_FILE, RAW_DIR, USER_DATA_BASE,
    load_db, save_db, get_ud_config,
    stok_dir as gtt_stok_dir, count_stok as gtt_count_stok, list_stok as gtt_list_stok,
    load_prompts, save_prompts, list_bahan_folders, list_bahan_images,
    escape_html, generate_stok_for_ud, merge_video_pair, GrokRateLimitError,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════
#  CONFIG
# ═══════════════════════════════════════════════════════════════
BOT_TOKEN = "YOUR_BOT_TOKEN_HERE"   # Ganti dengan token bot baru
ALLOWED_USER_IDS = []               # kosong = semua boleh

# Paths
GLOBAL_STOK_DIR     = os.path.join(APP_DIR, "global_stok")
GLOBAL_RAW_DIR      = os.path.join(APP_DIR, "global_raw")
SETTINGS_FILE       = os.path.join(APP_DIR, "autogen_settings.json")
SPREADSHEET_FILE    = os.path.join(APP_DIR, "autogen_log.xlsx")

# Brutal Bot references
BRUTAL_STOK_DIR     = os.path.join(APP_DIR, "brutal_stok")
BRUTAL_RAW_DIR      = os.path.join(APP_DIR, "brutal_stok_raw")
BRUTAL_SETTINGS_FILE = os.path.join(APP_DIR, "brutal_settings.json")
BRUTAL_MAX_STOK     = 50
BRUTAL_UD           = os.path.join(APP_DIR, "user_data", "brutal1")
BRUTAL_PORT         = "9260"
BRUTAL_MP3_DIR      = os.path.join(APP_DIR, "brutal_mp3")

# Grok config (shared UD for grok generation)
DEFAULT_GROK_UD     = os.path.join(USER_DATA_BASE, "autogen_grok")
DEFAULT_GROK_PORT   = "9280"

# Conversation states for text input
ASK_TAKE_COUNT = 1

# ═══════════════════════════════════════════════════════════════
#  STATE
# ═══════════════════════════════════════════════════════════════
daemon_task    = {}   # uid -> {stop, thread}
active_gen     = {}   # uid -> {stop, thread}

# ═══════════════════════════════════════════════════════════════
#  SETTINGS (persisted)
# ═══════════════════════════════════════════════════════════════
_DEFAULT_SETTINGS = {
    "priority": ["brutal", "ud_1", "ud_2"],
    "grok_ud": DEFAULT_GROK_UD,
    "grok_port": DEFAULT_GROK_PORT,
    "daemon_running": False,
    "pause_minutes_after_rate_limit": 30,
    "pause_minutes_between_targets": 2,
}

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return {**_DEFAULT_SETTINGS, **json.load(f)}
        except: pass
    return dict(_DEFAULT_SETTINGS)

def save_settings(cfg):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)

def is_allowed(uid):
    return not ALLOWED_USER_IDS or uid in ALLOWED_USER_IDS

# ═══════════════════════════════════════════════════════════════
#  SPREADSHEET LOGGING (7-day rotation)
# ═══════════════════════════════════════════════════════════════
_sheet_lock = threading.Lock()
LOG_RETENTION_DAYS = 7
LOG_ARCHIVE_DIR = os.path.join(APP_DIR, "autogen_log_archive")

def _init_spreadsheet():
    """Create spreadsheet if not exists."""
    if not os.path.exists(SPREADSHEET_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Generate Log"
        ws.append(["Nama Stok", "DateTime", "Status"])
        for col in range(1, 4):
            ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
        wb.save(SPREADSHEET_FILE)
        logger.info(f"📊 Spreadsheet dibuat: {SPREADSHEET_FILE}")

def _purge_old_entries():
    """Remove log entries older than 7 days. Called inside _sheet_lock."""
    if not os.path.exists(SPREADSHEET_FILE): return
    try:
        wb = load_workbook(SPREADSHEET_FILE)
        ws = wb.active
        cutoff = datetime.now() - timedelta(days=LOG_RETENTION_DAYS)
        rows_to_del = []
        for row_idx in range(2, ws.max_row + 1):
            dt_val = ws.cell(row=row_idx, column=2).value
            if not dt_val: continue
            try:
                entry_dt = datetime.strptime(str(dt_val), "%Y-%m-%d %H:%M:%S")
                if entry_dt < cutoff:
                    rows_to_del.append(row_idx)
            except: pass
        # Delete from bottom to top to keep indices valid
        for r in reversed(rows_to_del):
            ws.delete_rows(r)
        if rows_to_del:
            wb.save(SPREADSHEET_FILE)
            logger.info(f"🧹 Purged {len(rows_to_del)} old log entries (>7 hari)")
    except Exception as e:
        logger.error(f"Purge error: {e}")

def _get_oldest_entry_date():
    """Return datetime of the oldest entry, or None."""
    if not os.path.exists(SPREADSHEET_FILE): return None
    try:
        wb = load_workbook(SPREADSHEET_FILE)
        ws = wb.active
        if ws.max_row < 2: return None
        dt_val = ws.cell(row=2, column=2).value
        if dt_val:
            return datetime.strptime(str(dt_val), "%Y-%m-%d %H:%M:%S")
    except: pass
    return None

def _get_log_entry_count():
    """Return number of log entries (excluding header)."""
    if not os.path.exists(SPREADSHEET_FILE): return 0
    try:
        wb = load_workbook(SPREADSHEET_FILE)
        ws = wb.active
        return max(0, ws.max_row - 1)
    except: return 0

def _archive_and_reset_log():
    """Archive current spreadsheet and create fresh one. Called inside _sheet_lock."""
    if not os.path.exists(SPREADSHEET_FILE): return None
    os.makedirs(LOG_ARCHIVE_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_path = os.path.join(LOG_ARCHIVE_DIR, f"autogen_log_{ts}.xlsx")
    try:
        shutil.copy2(SPREADSHEET_FILE, archive_path)
        # Reset: create fresh spreadsheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Generate Log"
        ws.append(["Nama Stok", "DateTime", "Status"])
        for col in range(1, 4):
            ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
        wb.save(SPREADSHEET_FILE)
        logger.info(f"📊 Log diarsipkan → {archive_path}, spreadsheet direset")
        return archive_path
    except Exception as e:
        logger.error(f"Archive error: {e}")
        return None

def log_to_spreadsheet(nama_stok: str, status: str):
    """Log generate result to spreadsheet. Auto-purge >7 hari. Thread-safe."""
    with _sheet_lock:
        try:
            _init_spreadsheet()
            wb = load_workbook(SPREADSHEET_FILE)
            ws = wb.active
            ws.append([nama_stok, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), status])
            wb.save(SPREADSHEET_FILE)
            # Purge entries older than 7 days
            _purge_old_entries()
        except Exception as e:
            logger.error(f"Spreadsheet log error: {e}")

def check_and_send_weekly_log(bot, chat_id, main_loop):
    """
    Check if the oldest log entry is >= 7 days old.
    If yes, send the .xlsx to Telegram, archive it, and reset.
    Call this periodically from the daemon loop.
    """
    with _sheet_lock:
        oldest = _get_oldest_entry_date()
        count = _get_log_entry_count()
        if not oldest or count == 0:
            return False

        age = datetime.now() - oldest
        if age.days < LOG_RETENTION_DAYS:
            return False

        # Time to send and reset!
        logger.info(f"📊 Log sudah {age.days} hari, auto-send ke Telegram...")
        try:
            # Send file to Telegram
            async def _send_log():
                try:
                    with open(SPREADSHEET_FILE, "rb") as f:
                        await bot.send_document(
                            chat_id, document=f,
                            filename=f"autogen_log_weekly_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            caption=(
                                f"📊 <b>Log Mingguan Otomatis</b>\n\n"
                                f"Periode: <code>{oldest.strftime('%Y-%m-%d')}</code> s/d "
                                f"<code>{datetime.now().strftime('%Y-%m-%d')}</code>\n"
                                f"Total entry: <b>{count}</b>\n\n"
                                f"Log telah direset untuk 7 hari berikutnya."
                            ),
                            parse_mode=ParseMode.HTML
                        )
                except Exception as e:
                    await bot.send_message(chat_id,
                        f"⚠️ Gagal kirim log mingguan: {e}")

            future = asyncio.run_coroutine_threadsafe(_send_log(), main_loop)
            future.result(timeout=30)
        except Exception as e:
            logger.error(f"Auto-send log error: {e}")

        # Archive and reset
        _archive_and_reset_log()
        return True

# ═══════════════════════════════════════════════════════════════
#  STOK HELPERS
# ═══════════════════════════════════════════════════════════════

# — Brutal Bot stok —
def brutal_count_stok():
    if not os.path.isdir(BRUTAL_STOK_DIR): return 0
    return len([f for f in os.listdir(BRUTAL_STOK_DIR) if f.endswith(".mp4")])

def brutal_needed():
    return max(0, BRUTAL_MAX_STOK - brutal_count_stok())

def brutal_list_stok():
    if not os.path.isdir(BRUTAL_STOK_DIR): return []
    return sorted([os.path.join(BRUTAL_STOK_DIR, f)
                   for f in os.listdir(BRUTAL_STOK_DIR) if f.endswith(".mp4")],
                  key=os.path.getmtime)

def brutal_load_settings():
    if os.path.exists(BRUTAL_SETTINGS_FILE):
        try:
            with open(BRUTAL_SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except: pass
    return {}

# — Global stok —
def global_count_stok():
    os.makedirs(GLOBAL_STOK_DIR, exist_ok=True)
    return len([f for f in os.listdir(GLOBAL_STOK_DIR) if f.endswith(".mp4")])

def global_list_stok():
    os.makedirs(GLOBAL_STOK_DIR, exist_ok=True)
    return sorted([os.path.join(GLOBAL_STOK_DIR, f)
                   for f in os.listdir(GLOBAL_STOK_DIR) if f.endswith(".mp4")],
                  key=os.path.getmtime)

# — MP3 helpers (for brutal merge) —
def _brutal_list_mp3():
    os.makedirs(BRUTAL_MP3_DIR, exist_ok=True)
    return sorted([f for f in os.listdir(BRUTAL_MP3_DIR) if f.lower().endswith('.mp3')])

def _brutal_get_random_mp3():
    mp3s = _brutal_list_mp3()
    if not mp3s: return None
    return os.path.join(BRUTAL_MP3_DIR, random.choice(mp3s))

def _mute_and_add_mp3(video_path, log_fn=None):
    """Mute video + add random MP3 (same as brutal_bot)."""
    import subprocess
    mp3_path = _brutal_get_random_mp3()
    if not mp3_path:
        # Just mute
        tmp_out = video_path + ".muted.mp4"
        cmd = ["ffmpeg", "-y", "-i", video_path, "-an", "-c:v", "copy", tmp_out]
        try:
            r = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
            if r.returncode == 0 and os.path.exists(tmp_out) and os.path.getsize(tmp_out) > 0:
                os.replace(tmp_out, video_path)
                return True
        except: pass
        try:
            if os.path.exists(tmp_out): os.remove(tmp_out)
        except: pass
        return False

    tmp_out = video_path + ".audio.mp4"
    cmd = [
        "ffmpeg", "-y", "-i", video_path,
        "-stream_loop", "-1", "-i", mp3_path,
        "-c:v", "copy", "-c:a", "aac", "-b:a", "128k",
        "-map", "0:v:0", "-map", "1:a:0", "-shortest", tmp_out
    ]
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
        if r.returncode == 0 and os.path.exists(tmp_out) and os.path.getsize(tmp_out) > 0:
            os.replace(tmp_out, video_path)
            return True
    except: pass
    try:
        if os.path.exists(tmp_out): os.remove(tmp_out)
    except: pass
    return False

def brutal_merge_video_pair(vid1, vid2, output_dir, log_fn=None):
    """Merge 2 video for Brutal Bot (concat + MP3 replace)."""
    os.makedirs(output_dir, exist_ok=True)
    existing = glob.glob(os.path.join(output_dir, "*.mp4"))
    nums = []
    for f in existing:
        m = re.fullmatch(r'(\d+)\.mp4', os.path.basename(f))
        if m: nums.append(int(m.group(1)))
    next_num = (max(nums) + 1) if nums else 1
    out_path = os.path.join(output_dir, f"{next_num}.mp4")
    list_file = os.path.join(output_dir, f"_mlist_{next_num}.txt")
    try:
        with open(list_file, "w", encoding="utf-8") as lf:
            lf.write(f"file '{vid1}'\nfile '{vid2}'\n")
        import subprocess
        cmd = ["ffmpeg","-y","-f","concat","-safe","0","-i",list_file,"-c","copy",out_path]
        if log_fn: log_fn(f"🎬 Merge: {os.path.basename(vid1)} + {os.path.basename(vid2)} → {next_num}.mp4")
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if r.returncode == 0 and os.path.exists(out_path) and os.path.getsize(out_path) > 0:
            if log_fn: log_fn(f"✅ Merged: {next_num}.mp4")
            _mute_and_add_mp3(out_path, log_fn)
            return out_path
        if log_fn: log_fn(f"❌ Merge gagal")
        return None
    except Exception as e:
        if log_fn: log_fn(f"❌ Merge error: {e}")
        return None
    finally:
        try:
            if os.path.exists(list_file): os.remove(list_file)
        except: pass

# ═══════════════════════════════════════════════════════════════
#  TARGET RESOLUTION — Get prompt/bahan/dir for each target
# ═══════════════════════════════════════════════════════════════
def resolve_target(target_name):
    """
    Return dict: {name, prompt_text, bahan_folder, stok_dir, raw_dir, needed,
                  grok_ud, grok_port, merge_func, max_stok, current_stok}
    or None if target not configured.
    """
    settings = load_settings()
    prompts = load_prompts()

    if target_name == "brutal":
        bs = brutal_load_settings()
        prompt_name = bs.get("prompt_name", "")
        folder_name = bs.get("folder_name", "")
        prompt_text = prompts.get(prompt_name, "")
        if not prompt_text or not folder_name:
            return None
        return {
            "name": "Brutal Bot",
            "ud_label": "Brutal",
            "prompt_text": prompt_text,
            "bahan_folder": folder_name,
            "stok_dir": BRUTAL_STOK_DIR,
            "raw_dir": BRUTAL_RAW_DIR,
            "needed": brutal_needed(),
            "max_stok": BRUTAL_MAX_STOK,
            "current_stok": brutal_count_stok(),
            "merge_func": brutal_merge_video_pair,
        }

    # UD targets: ud_1, ud_2, ...
    m = re.match(r"ud_(\d+)", target_name)
    if m:
        ud_num = int(m.group(1))
        db = load_db()
        cfg = get_ud_config(db, ud_num)
        prompt_name = cfg.get("prompt_name", "")
        bahan_folder = cfg.get("bahan_folder", "")
        prompt_text = prompts.get(prompt_name, "")
        if not prompt_text or not bahan_folder:
            return None
        batch_size = cfg.get("batch_size", 30)
        current = gtt_count_stok(ud_num)
        return {
            "name": f"UD {ud_num}",
            "ud_label": str(ud_num),
            "prompt_text": prompt_text,
            "bahan_folder": bahan_folder,
            "stok_dir": gtt_stok_dir(ud_num),
            "raw_dir": RAW_DIR,
            "needed": max(0, batch_size - current),
            "max_stok": batch_size,
            "current_stok": current,
            "merge_func": None,  # default gtt_core merge
        }

    if target_name == "global":
        # For global, use the FIRST configured target's prompt/bahan as default
        priority = settings.get("priority", ["brutal"])
        for tgt in priority:
            resolved = resolve_target(tgt)
            if resolved:
                return {
                    "name": "Stok Global",
                    "ud_label": "Global",
                    "prompt_text": resolved["prompt_text"],
                    "bahan_folder": resolved["bahan_folder"],
                    "stok_dir": GLOBAL_STOK_DIR,
                    "raw_dir": GLOBAL_RAW_DIR,
                    "needed": 10,  # always generate 10 at a time for global
                    "max_stok": 9999,
                    "current_stok": global_count_stok(),
                    "merge_func": resolved.get("merge_func"),
                }
        return None

    return None

# ═══════════════════════════════════════════════════════════════
#  24H DAEMON — Continuous generate loop
# ═══════════════════════════════════════════════════════════════
def run_24h_daemon(uid, chat_id, bot, main_loop, stop_event):
    """Main daemon loop: generate continuously based on priority."""
    def send(text):
        asyncio.run_coroutine_threadsafe(
            bot.send_message(chat_id, text, parse_mode=ParseMode.HTML), main_loop)

    settings = load_settings()
    grok_ud = settings.get("grok_ud", DEFAULT_GROK_UD)
    grok_port = settings.get("grok_port", DEFAULT_GROK_PORT)
    pause_rate = settings.get("pause_minutes_after_rate_limit", 30)
    pause_between = settings.get("pause_minutes_between_targets", 2)

    send(f"<b>🤖 Auto Generator 24-Jam Aktif!</b>\n\n"
         f"Prioritas: <code>{' → '.join(settings.get('priority', []))}</code>\n"
         f"Grok UD: <code>{grok_ud}</code>\n"
         f"Grok Port: <code>{grok_port}</code>\n\n"
         f"Jika semua stok penuh → Generate ke <b>Stok Global</b>")

    cycle = 0
    while not stop_event.is_set():
        cycle += 1

        # ── Cek & kirim log mingguan otomatis ──
        try:
            check_and_send_weekly_log(bot, chat_id, main_loop)
        except Exception as e:
            logger.error(f"Weekly log check error: {e}")

        settings = load_settings()
        priority = settings.get("priority", ["brutal", "ud_1", "ud_2"])
        grok_ud = settings.get("grok_ud", DEFAULT_GROK_UD)
        grok_port = settings.get("grok_port", DEFAULT_GROK_PORT)

        target_found = False
        all_full = True

        for target_name in priority:
            if stop_event.is_set(): break
            info = resolve_target(target_name)
            if not info:
                continue

            if info["needed"] <= 0:
                continue  # stok penuh untuk target ini

            all_full = False
            target_found = True
            needed = info["needed"]
            send(f"<b>🎯 [{info['name']}]</b> Generate {needed} video\n"
                 f"Stok: {info['current_stok']}/{info['max_stok']}\n"
                 f"Prompt: <code>{escape_html(info['prompt_text'][:60])}...</code>")

            # Generate with logging
            gen_success = False
            try:
                os.makedirs(info["stok_dir"], exist_ok=True)
                os.makedirs(info["raw_dir"], exist_ok=True)

                log_lines = []
                log_lock = threading.Lock()

                def gen_log(msg):
                    with log_lock:
                        log_lines.append(f"<code>[{datetime.now().strftime('%H:%M:%S')}]</code> {escape_html(str(msg))}")
                        if len(log_lines) > 25: log_lines.pop(0)

                merged = generate_stok_for_ud(
                    ud_num=info["ud_label"],
                    needed=needed,
                    prompt_text=info["prompt_text"],
                    bahan_folder=info["bahan_folder"],
                    grok_ud=grok_ud,
                    grok_port=grok_port,
                    log_fn=gen_log,
                    stop_event=stop_event,
                    out_dir=info["stok_dir"],
                    raw_dir=info["raw_dir"],
                    merge_func=info.get("merge_func"),
                )
                gen_success = True
                log_to_spreadsheet(info["name"], "Berhasil")
                send(f"<b>✅ [{info['name']}]</b> Generate selesai!\n"
                     f"Merged: {merged} video")

            except GrokRateLimitError:
                log_to_spreadsheet(info["name"], "Gagal - Rate Limit")
                send(f"🚫 <b>RATE LIMIT!</b> [{info['name']}]\n"
                     f"Pause {pause_rate} menit sebelum lanjut...")
                for _ in range(int(pause_rate * 60 / 5)):
                    if stop_event.is_set(): break
                    time.sleep(5)
                continue

            except Exception as e:
                log_to_spreadsheet(info["name"], f"Gagal - {type(e).__name__}")
                send(f"❌ <b>[{info['name']}] Error:</b> {type(e).__name__}\n"
                     f"<code>{escape_html(str(e)[:150])}</code>\n"
                     f"Retry dalam 60 detik...")
                for _ in range(12):
                    if stop_event.is_set(): break
                    time.sleep(5)
                continue

            # Pause between targets
            if not stop_event.is_set() and pause_between > 0:
                send(f"⏸ Pause {pause_between} menit sebelum target berikutnya...")
                for _ in range(int(pause_between * 60 / 5)):
                    if stop_event.is_set(): break
                    time.sleep(5)

        if stop_event.is_set(): break

        # Jika semua stok penuh → generate ke global
        if all_full or not target_found:
            global_info = resolve_target("global")
            if global_info:
                send(f"<b>🌐 Semua stok penuh!</b>\n"
                     f"Generate ke <b>Stok Global</b> ({global_count_stok()} video saat ini)")
                try:
                    os.makedirs(GLOBAL_STOK_DIR, exist_ok=True)
                    os.makedirs(GLOBAL_RAW_DIR, exist_ok=True)

                    def gl_log(msg):
                        pass  # silent for global

                    merged = generate_stok_for_ud(
                        ud_num="Global",
                        needed=10,
                        prompt_text=global_info["prompt_text"],
                        bahan_folder=global_info["bahan_folder"],
                        grok_ud=grok_ud,
                        grok_port=grok_port,
                        log_fn=gl_log,
                        stop_event=stop_event,
                        out_dir=GLOBAL_STOK_DIR,
                        raw_dir=GLOBAL_RAW_DIR,
                        merge_func=global_info.get("merge_func"),
                    )
                    log_to_spreadsheet("Stok Global", "Berhasil")
                    send(f"<b>🌐 Stok Global:</b> +{merged} video (total: {global_count_stok()})")
                except GrokRateLimitError:
                    log_to_spreadsheet("Stok Global", "Gagal - Rate Limit")
                    send(f"🚫 <b>RATE LIMIT!</b> Pause {pause_rate} menit...")
                    for _ in range(int(pause_rate * 60 / 5)):
                        if stop_event.is_set(): break
                        time.sleep(5)
                except Exception as e:
                    log_to_spreadsheet("Stok Global", f"Gagal - {type(e).__name__}")
                    send(f"❌ Global error: {escape_html(str(e)[:100])}")
                    time.sleep(60)
            else:
                send("⚠️ Tidak ada target yang terkonfigurasi! Menunggu 5 menit...")
                for _ in range(60):
                    if stop_event.is_set(): break
                    time.sleep(5)

    daemon_task.pop(uid, None)
    send("<b>🛑 Auto Generator 24-Jam dihentikan.</b>")


# ═══════════════════════════════════════════════════════════════
#  MOVE GLOBAL STOK TO TARGET
# ═══════════════════════════════════════════════════════════════
def move_global_to_target(target_name, count):
    """Move 'count' videos from global stok to target stok. Return actual moved count."""
    info = resolve_target(target_name)
    if not info: return 0

    global_files = global_list_stok()
    if not global_files: return 0

    to_move = min(count, len(global_files), info["needed"])
    moved = 0
    for fp in global_files[:to_move]:
        dest = os.path.join(info["stok_dir"], os.path.basename(fp))
        try:
            shutil.move(fp, dest)
            moved += 1
        except: pass
    return moved


# ═══════════════════════════════════════════════════════════════
#  TELEGRAM — MENU
# ═══════════════════════════════════════════════════════════════
def main_menu_kb(uid=None):
    is_daemon = bool(uid and daemon_task.get(uid))
    settings = load_settings()
    priority = settings.get("priority", [])

    rows = [
        [InlineKeyboardButton("📊 Status Stok", callback_data="status"),
         InlineKeyboardButton("⚙️ Prioritas", callback_data="priority_menu")],
        [InlineKeyboardButton("📝 Kelola Prompt", callback_data="prompt_menu"),
         InlineKeyboardButton("📁 Kelola Bahan", callback_data="bahan_menu")],
        [InlineKeyboardButton(f"📦 Ambil Stok Global ({global_count_stok()})", callback_data="take_global")],
        [InlineKeyboardButton("📊 Lihat Log", callback_data="view_log"),
         InlineKeyboardButton("📥 Download .xlsx", callback_data="download_log")],
    ]

    if is_daemon:
        rows.append([InlineKeyboardButton("🛑 Stop 24-Jam", callback_data="stop_daemon")])
    else:
        rows.append([InlineKeyboardButton("🚀 Start 24-Jam", callback_data="start_daemon")])

    rows.append([InlineKeyboardButton("🔄 Refresh", callback_data="refresh")])
    return InlineKeyboardMarkup(rows)


def status_text():
    settings = load_settings()
    priority = settings.get("priority", [])
    prompts = load_prompts()

    lines = ["<b>🤖 Grok Auto Generator</b>\n"]

    # Daemon status
    is_running = bool(daemon_task)
    lines.append(f"Status: <b>{'🟢 AKTIF' if is_running else '🔴 MATI'}</b>")
    lines.append(f"Prioritas: <code>{' → '.join(priority)}</code>\n")

    # Per-target stok
    for tgt in priority:
        info = resolve_target(tgt)
        if info:
            pct = int(info['current_stok'] / info['max_stok'] * 100) if info['max_stok'] > 0 else 0
            bar = "█" * (pct // 10) + "░" * (10 - pct // 10)
            lines.append(
                f"<b>{info['name']}:</b> {info['current_stok']}/{info['max_stok']} "
                f"[{bar}] {pct}%")
        else:
            lines.append(f"<b>{tgt}:</b> <i>belum dikonfigurasi</i>")

    # Global stok
    gc = global_count_stok()
    lines.append(f"\n<b>🌐 Stok Global:</b> {gc} video")

    # Grok config
    lines.append(f"\nGrok UD: <code>{settings.get('grok_ud', '')}</code>")
    lines.append(f"Grok Port: <code>{settings.get('grok_port', '')}</code>")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════
#  TELEGRAM — HANDLERS
# ═══════════════════════════════════════════════════════════════
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not is_allowed(uid): return
    await update.message.reply_text(status_text(), parse_mode=ParseMode.HTML, reply_markup=main_menu_kb(uid))

async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "<b>🤖 Grok Auto Generator Bot</b>\n\n"
        "<b>Perintah:</b>\n"
        "/start - Menu utama\n"
        "/help - Panduan\n"
        "/stop - Stop semua proses\n\n"
        "<b>Konfigurasi:</b>\n"
        "<code>/set grok_ud FOLDER</code> - Grok user data\n"
        "<code>/set grok_port PORT</code> - Grok port\n"
        "<code>/set pause_rate MENIT</code> - Pause setelah rate limit\n"
        "<code>/set pause_between MENIT</code> - Pause antar target\n\n"
        "<b>Flow:</b>\n"
        "1. Pastikan Brutal Bot & GTT Bot sudah dikonfigurasi\n"
        "2. Atur prioritas via menu ⚙️ Prioritas\n"
        "3. Klik 🚀 Start 24-Jam\n"
        "4. Bot akan generate terus-menerus\n"
        "5. Gunakan 📦 Ambil Stok Global untuk download video",
        parse_mode=ParseMode.HTML)

async def cmd_stop(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    stopped = []
    t = daemon_task.get(uid)
    if t: t["stop"].set(); daemon_task.pop(uid, None); stopped.append("24-Jam Daemon")
    t = active_gen.get(uid)
    if t: t["stop"].set(); active_gen.pop(uid, None); stopped.append("Generate")
    if stopped:
        await update.message.reply_text(f"Dihentikan: {', '.join(stopped)}")
    else:
        await update.message.reply_text("Tidak ada proses berjalan.")

async def cmd_log(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Langsung kirim file .xlsx log ke chat."""
    uid = update.effective_user.id
    if not is_allowed(uid): return
    if not os.path.exists(SPREADSHEET_FILE):
        _init_spreadsheet()
    try:
        with open(SPREADSHEET_FILE, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename="autogen_log.xlsx",
                caption="📊 Spreadsheet Log Generate\n"
                        f"File: <code>{SPREADSHEET_FILE}</code>",
                parse_mode=ParseMode.HTML
            )
    except Exception as e:
        await update.message.reply_text(f"❌ Gagal kirim file: {e}")

async def cmd_set(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not is_allowed(uid): return
    raw = update.message.text.strip(); args = raw.split(None, 2)
    if len(args) < 3:
        await update.message.reply_text(
            "<b>Format /set:</b>\n"
            "<code>/set grok_ud FOLDER</code>\n"
            "<code>/set grok_port PORT</code>\n"
            "<code>/set pause_rate 30</code> - menit setelah rate limit\n"
            "<code>/set pause_between 2</code> - menit antar target",
            parse_mode=ParseMode.HTML); return
    sub = args[1].lower(); val = args[2].strip()
    s = load_settings()
    if sub == "grok_ud":
        if ":" not in val and not val.startswith("\\"):
            val = os.path.join(USER_DATA_BASE, val)
        s["grok_ud"] = val
    elif sub == "grok_port":
        s["grok_port"] = val
    elif sub == "pause_rate":
        try: s["pause_minutes_after_rate_limit"] = int(val)
        except: pass
    elif sub == "pause_between":
        try: s["pause_minutes_between_targets"] = int(val)
        except: pass
    else:
        await update.message.reply_text("Sub-command tidak dikenal."); return
    save_settings(s)
    await update.message.reply_text(f"<code>{sub}</code> = <code>{escape_html(str(val)[:100])}</code>",
                                     parse_mode=ParseMode.HTML)


# ═══════════════════════════════════════════════════════════════
#  TELEGRAM — BUTTON HANDLER
# ═══════════════════════════════════════════════════════════════
async def button_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    uid = q.from_user.id
    if not is_allowed(uid): return
    data = q.data; chat_id = q.message.chat_id
    bot = ctx.bot; main_loop = asyncio.get_event_loop()
    back_kb = InlineKeyboardMarkup([[InlineKeyboardButton("Kembali", callback_data="refresh")]])

    # ── REFRESH/STATUS ──
    if data in ("refresh", "status"):
        await q.edit_message_text(status_text(), parse_mode=ParseMode.HTML, reply_markup=main_menu_kb(uid))
        return

    # ── START DAEMON ──
    if data == "start_daemon":
        if daemon_task.get(uid):
            await q.edit_message_text("24-Jam sudah berjalan!", reply_markup=main_menu_kb(uid)); return
        stop_evt = threading.Event()
        t = threading.Thread(target=run_24h_daemon, args=(uid, chat_id, bot, main_loop, stop_evt), daemon=True)
        daemon_task[uid] = {"stop": stop_evt, "thread": t}; t.start()
        await q.edit_message_text(
            "<b>🚀 Auto Generator 24-Jam Aktif!</b>\nTekan 🛑 Stop untuk menghentikan.",
            parse_mode=ParseMode.HTML, reply_markup=main_menu_kb(uid)); return

    # ── STOP DAEMON ──
    if data == "stop_daemon":
        t = daemon_task.get(uid)
        if t: t["stop"].set(); daemon_task.pop(uid, None)
        await q.edit_message_text("<b>🛑 Generator dihentikan.</b>\n\n" + status_text(),
                                   parse_mode=ParseMode.HTML, reply_markup=main_menu_kb(uid)); return

    # ── PRIORITY MENU ──
    if data == "priority_menu":
        settings = load_settings()
        priority = settings.get("priority", [])
        rows = []
        for i, tgt in enumerate(priority):
            info = resolve_target(tgt)
            label = info["name"] if info else tgt
            btns = []
            if i > 0:
                btns.append(InlineKeyboardButton("⬆️", callback_data=f"prio_up|{i}"))
            if i < len(priority) - 1:
                btns.append(InlineKeyboardButton("⬇️", callback_data=f"prio_down|{i}"))
            btns.append(InlineKeyboardButton(f"{i+1}. {label}", callback_data=f"prio_info|{tgt}"))
            btns.append(InlineKeyboardButton("❌", callback_data=f"prio_del|{i}"))
            rows.append(btns)
        rows.append([InlineKeyboardButton("+ Tambah Target", callback_data="prio_add")])
        rows.append([InlineKeyboardButton("Kembali", callback_data="refresh")])
        text = (f"<b>⚙️ Prioritas Generate</b>\n\n"
                f"Urutan: <code>{' → '.join(priority)}</code>\n\n"
                f"Target pertama paling prioritas.\n"
                f"Jika semua penuh → masuk <b>Stok Global</b>.")
        await q.edit_message_text(text, parse_mode=ParseMode.HTML, reply_markup=InlineKeyboardMarkup(rows)); return

    if data.startswith("prio_up|"):
        idx = int(data.split("|")[1])
        settings = load_settings()
        p = settings.get("priority", [])
        if 0 < idx < len(p):
            p[idx], p[idx-1] = p[idx-1], p[idx]
            settings["priority"] = p; save_settings(settings)
        # Re-show priority menu
        await button_handler.__wrapped__(update, ctx) if hasattr(button_handler, '__wrapped__') else None
        # Simulate re-click priority_menu
        q.data = "priority_menu"
        await button_handler(update, ctx); return

    if data.startswith("prio_down|"):
        idx = int(data.split("|")[1])
        settings = load_settings()
        p = settings.get("priority", [])
        if 0 <= idx < len(p) - 1:
            p[idx], p[idx+1] = p[idx+1], p[idx]
            settings["priority"] = p; save_settings(settings)
        q.data = "priority_menu"
        await button_handler(update, ctx); return

    if data.startswith("prio_del|"):
        idx = int(data.split("|")[1])
        settings = load_settings()
        p = settings.get("priority", [])
        if 0 <= idx < len(p):
            removed = p.pop(idx)
            settings["priority"] = p; save_settings(settings)
        q.data = "priority_menu"
        await button_handler(update, ctx); return

    if data == "prio_add":
        settings = load_settings()
        p = settings.get("priority", [])
        # Show available targets not yet in priority
        available = []
        if "brutal" not in p:
            available.append(("brutal", "Brutal Bot"))
        db = load_db()
        for ud_num in db.get("active_ud", [1, 2]):
            key = f"ud_{ud_num}"
            if key not in p:
                available.append((key, f"UD {ud_num}"))
        # Also show potential UDs 1-7
        for i in range(1, 8):
            key = f"ud_{i}"
            if key not in p and (key, f"UD {i}") not in available:
                available.append((key, f"UD {i}"))

        if not available:
            await q.edit_message_text("Semua target sudah ditambahkan!", reply_markup=back_kb); return

        rows = []
        for key, label in available:
            rows.append([InlineKeyboardButton(label, callback_data=f"prio_add_do|{key}")])
        rows.append([InlineKeyboardButton("Kembali", callback_data="priority_menu")])
        await q.edit_message_text("<b>Pilih target untuk ditambahkan:</b>",
                                   parse_mode=ParseMode.HTML, reply_markup=InlineKeyboardMarkup(rows)); return

    if data.startswith("prio_add_do|"):
        key = data.split("|")[1]
        settings = load_settings()
        p = settings.get("priority", [])
        if key not in p:
            p.append(key)
            settings["priority"] = p; save_settings(settings)
        q.data = "priority_menu"
        await button_handler(update, ctx); return

    # ── PROMPT MENU ──
    if data == "prompt_menu":
        prompts = load_prompts()
        rows = []
        for name in prompts:
            rows.append([InlineKeyboardButton(name, callback_data=f"prompt_view|{name}")])
        rows.append([InlineKeyboardButton("+ Tambah Prompt", callback_data="prompt_add")])
        rows.append([InlineKeyboardButton("Kembali", callback_data="refresh")])
        await q.edit_message_text(f"<b>📝 Kelola Prompt</b>\n{len(prompts)} prompt",
                                   parse_mode=ParseMode.HTML, reply_markup=InlineKeyboardMarkup(rows)); return

    if data.startswith("prompt_view|"):
        name = data.split("|", 1)[1]
        prompts = load_prompts()
        text_val = prompts.get(name, "(tidak ditemukan)")
        rows = [
            [InlineKeyboardButton("Hapus", callback_data=f"prompt_del|{name}")],
            [InlineKeyboardButton("Kembali", callback_data="prompt_menu")]
        ]
        await q.edit_message_text(
            f"<b>Prompt: {escape_html(name)}</b>\n\n<code>{escape_html(text_val[:500])}</code>",
            parse_mode=ParseMode.HTML, reply_markup=InlineKeyboardMarkup(rows)); return

    if data == "prompt_add":
        ctx.user_data["waiting_for"] = "prompt_name"
        await bot.send_message(chat_id, "Kirim <b>nama</b> prompt baru:", parse_mode=ParseMode.HTML); return

    if data.startswith("prompt_del|"):
        name = data.split("|", 1)[1]
        prompts = load_prompts()
        prompts.pop(name, None); save_prompts(prompts)
        await q.edit_message_text(f"Prompt <code>{escape_html(name)}</code> dihapus!",
                                   parse_mode=ParseMode.HTML, reply_markup=back_kb); return

    # ── BAHAN MENU ──
    if data == "bahan_menu":
        folders = list_bahan_folders()
        rows = []
        for f in folders:
            imgs = list_bahan_images(f)
            rows.append([InlineKeyboardButton(f"{f} ({len(imgs)})", callback_data=f"bahan_view|{f}")])
        rows.append([InlineKeyboardButton("+ Tambah Folder", callback_data="bahan_add")])
        rows.append([InlineKeyboardButton("Kembali", callback_data="refresh")])
        await q.edit_message_text(f"<b>📁 Kelola Bahan</b>\n{len(folders)} folder",
                                   parse_mode=ParseMode.HTML, reply_markup=InlineKeyboardMarkup(rows)); return

    if data.startswith("bahan_view|"):
        folder = data.split("|", 1)[1]
        imgs = list_bahan_images(folder)
        text = f"<b>Bahan: {escape_html(folder)}</b>\n{len(imgs)} gambar\n\n"
        for i, img in enumerate(imgs[:20]):
            text += f"  {i+1}. <code>{os.path.basename(img)}</code>\n"
        if len(imgs) > 20: text += f"  ... +{len(imgs)-20} lagi\n"
        text += "\nKirim foto untuk menambah gambar."
        ctx.user_data["waiting_for"] = f"bahan_photo|{folder}"
        rows = [
            [InlineKeyboardButton("Hapus Folder", callback_data=f"bahan_del|{folder}")],
            [InlineKeyboardButton("Kembali", callback_data="bahan_menu")]
        ]
        await q.edit_message_text(text, parse_mode=ParseMode.HTML, reply_markup=InlineKeyboardMarkup(rows)); return

    if data == "bahan_add":
        ctx.user_data["waiting_for"] = "bahan_new_folder"
        await bot.send_message(chat_id, "Kirim nama folder baru:"); return

    if data.startswith("bahan_del|"):
        folder = data.split("|", 1)[1]
        path = os.path.join(BAHAN_DIR, folder)
        try:
            if os.path.isdir(path): shutil.rmtree(path)
            await q.edit_message_text(f"Folder <code>{escape_html(folder)}</code> dihapus!",
                                       parse_mode=ParseMode.HTML, reply_markup=back_kb)
        except Exception as e:
            await q.edit_message_text(f"Gagal hapus: {e}", reply_markup=back_kb)
        return

    # ── TAKE GLOBAL STOK ──
    if data == "take_global":
        gc = global_count_stok()
        if gc == 0:
            await q.edit_message_text("📦 <b>Stok Global kosong!</b>\nTunggu hingga generator mengisi stok global.",
                                       parse_mode=ParseMode.HTML, reply_markup=main_menu_kb(uid)); return
        ctx.user_data["waiting_for"] = "take_global_count"
        await q.edit_message_text(
            f"📦 <b>Stok Global Tersedia: {gc} video</b>\n\n"
            f"Kirim jumlah video yang ingin diambil (angka):",
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Batal", callback_data="refresh")]]))
        return

    # ── VIEW LOG ──
    if data == "view_log":
        if not os.path.exists(SPREADSHEET_FILE):
            await q.edit_message_text("📊 <b>Belum ada log generate.</b>",
                                       parse_mode=ParseMode.HTML, reply_markup=main_menu_kb(uid)); return
        try:
            wb = load_workbook(SPREADSHEET_FILE)
            ws = wb.active
            rows_data = list(ws.iter_rows(values_only=True))
            if len(rows_data) <= 1:
                await q.edit_message_text("📊 <b>Log masih kosong.</b>",
                                           parse_mode=ParseMode.HTML, reply_markup=main_menu_kb(uid)); return
            # Show last 20 entries
            entries = rows_data[1:][-20:]
            lines = ["<b>📊 Log Generate (terakhir 20)</b>\n"]
            for nama, dt, status in entries:
                emoji = "✅" if status and "Berhasil" in str(status) else "❌"
                lines.append(f"{emoji} <code>{dt}</code> | {escape_html(str(nama))} | {escape_html(str(status))}")

            # Stats
            total = len(rows_data) - 1
            success = sum(1 for r in rows_data[1:] if r[2] and "Berhasil" in str(r[2]))
            fail = total - success
            lines.append(f"\n<b>Total:</b> {total} | ✅ {success} | ❌ {fail}")

            text = "\n".join(lines)
            # Send spreadsheet file too
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("📥 Download Spreadsheet", callback_data="download_log")],
                [InlineKeyboardButton("Kembali", callback_data="refresh")]
            ])
            await q.edit_message_text(text[:4096], parse_mode=ParseMode.HTML, reply_markup=kb)
        except Exception as e:
            await q.edit_message_text(f"Error membaca log: {e}", reply_markup=main_menu_kb(uid))
        return

    if data == "download_log":
        if os.path.exists(SPREADSHEET_FILE):
            try:
                await bot.send_document(chat_id, document=open(SPREADSHEET_FILE, "rb"),
                                         filename="autogen_log.xlsx",
                                         caption="📊 Spreadsheet Log Generate")
            except Exception as e:
                await bot.send_message(chat_id, f"Gagal kirim file: {e}")
        else:
            await bot.send_message(chat_id, "File log tidak ditemukan.")
        return

    if data.startswith("prio_info|"):
        tgt = data.split("|")[1]
        info = resolve_target(tgt)
        if info:
            text = (f"<b>Info: {info['name']}</b>\n\n"
                    f"Stok: <b>{info['current_stok']}/{info['max_stok']}</b>\n"
                    f"Needed: <b>{info['needed']}</b>\n"
                    f"Prompt: <code>{escape_html(info['prompt_text'][:80])}...</code>\n"
                    f"Bahan: <code>{info['bahan_folder']}</code>")
        else:
            text = f"<b>{tgt}</b>: belum dikonfigurasi"
        await q.edit_message_text(text, parse_mode=ParseMode.HTML, reply_markup=back_kb); return


# ═══════════════════════════════════════════════════════════════
#  TEXT & PHOTO HANDLERS
# ═══════════════════════════════════════════════════════════════
async def text_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.text: return
    uid = update.effective_user.id
    if not is_allowed(uid): return
    waiting = ctx.user_data.get("waiting_for", "")
    text = update.message.text.strip()

    if waiting == "bahan_new_folder":
        safe = re.sub(r'[^\w\-]', '_', text)
        os.makedirs(os.path.join(BAHAN_DIR, safe), exist_ok=True)
        ctx.user_data["waiting_for"] = f"bahan_photo|{safe}"
        await update.message.reply_text(
            f"Folder <code>{escape_html(safe)}</code> dibuat!\nKirim foto untuk menambah.",
            parse_mode=ParseMode.HTML)
        return

    if waiting == "prompt_name":
        ctx.user_data["new_prompt_name"] = text
        ctx.user_data["waiting_for"] = "prompt_text"
        await update.message.reply_text(
            f"Nama: <code>{escape_html(text)}</code>\nSekarang kirim <b>teks prompt</b>:",
            parse_mode=ParseMode.HTML)
        return

    if waiting == "prompt_text":
        name = ctx.user_data.get("new_prompt_name", "unnamed")
        prompts = load_prompts()
        prompts[name] = text
        save_prompts(prompts)
        ctx.user_data["waiting_for"] = ""
        await update.message.reply_text(f"Prompt <code>{escape_html(name)}</code> disimpan!",
                                         parse_mode=ParseMode.HTML)
        return

    if waiting == "take_global_count":
        ctx.user_data["waiting_for"] = ""
        try:
            count = int(text)
        except ValueError:
            await update.message.reply_text("⚠️ Kirim angka yang valid!"); return

        gc = global_count_stok()
        if count <= 0:
            await update.message.reply_text("⚠️ Jumlah harus > 0"); return
        if count > gc:
            count = gc
            await update.message.reply_text(f"📦 Hanya {gc} tersedia, mengirim semua...")

        global_files = global_list_stok()[:count]
        await update.message.reply_text(f"📤 Mengirim {len(global_files)} video...")

        sent = 0
        for fp in global_files:
            try:
                await update.message.reply_video(
                    video=open(fp, "rb"),
                    caption=f"📦 Global Stok #{sent+1}",
                    supports_streaming=True
                )
                sent += 1
                # Remove from global stok after sending
                try: os.remove(fp)
                except: pass
            except Exception as e:
                logger.error(f"Send video error: {e}")
                try:
                    await update.message.reply_document(
                        document=open(fp, "rb"),
                        caption=f"📦 Global Stok #{sent+1}"
                    )
                    sent += 1
                    try: os.remove(fp)
                    except: pass
                except:
                    await update.message.reply_text(f"❌ Gagal kirim: {os.path.basename(fp)}")

        await update.message.reply_text(
            f"✅ <b>{sent}/{count} video dikirim!</b>\nSisa stok global: {global_count_stok()}",
            parse_mode=ParseMode.HTML, reply_markup=main_menu_kb(uid))
        return


async def photo_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.photo: return
    uid = update.effective_user.id
    if not is_allowed(uid): return
    waiting = ctx.user_data.get("waiting_for", "")
    if not waiting.startswith("bahan_photo|"): return
    folder = waiting.split("|", 1)[1]
    folder_path = os.path.join(BAHAN_DIR, folder)
    os.makedirs(folder_path, exist_ok=True)
    photo = update.message.photo[-1]
    file = await photo.get_file()
    filename = f"img_{int(datetime.now().timestamp())}_{random.randint(100,999)}.jpg"
    filepath = os.path.join(folder_path, filename)
    await file.download_to_drive(filepath)
    count = len(list_bahan_images(folder))
    await update.message.reply_text(
        f"Gambar disimpan ke <code>{escape_html(folder)}</code> ({count} gambar)",
        parse_mode=ParseMode.HTML)


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════
async def post_init(application):
    await application.bot.set_my_commands([
        BotCommand("start", "Menu utama"),
        BotCommand("set", "Konfigurasi"),
        BotCommand("log", "Download log .xlsx"),
        BotCommand("help", "Panduan"),
        BotCommand("stop", "Stop proses"),
    ])

def main():
    if BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
        print("ERROR: Set BOT_TOKEN dulu di grok_auto_gen_bot.py!")
        return

    # Ensure directories
    for d in [GLOBAL_STOK_DIR, GLOBAL_RAW_DIR, BRUTAL_STOK_DIR, BRUTAL_RAW_DIR]:
        os.makedirs(d, exist_ok=True)

    # Init spreadsheet
    _init_spreadsheet()

    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("set", cmd_set))
    app.add_handler(CommandHandler("log", cmd_log))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("stop", cmd_stop))
    app.add_handler(MessageHandler(filters.PHOTO, photo_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))
    app.add_handler(CallbackQueryHandler(button_handler))
    print("🤖 Grok Auto Generator Bot is running...")
    app.run_polling()

if __name__ == "__main__":
    main()
